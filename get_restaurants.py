import requests
from openpyxl import Workbook
URL = "https://api.yelp.com/v3/businesses/search"
API_KEY = "_9wB6XJ3OidqsnabVaUdwnBmx7aG7WqOrB_wvEh99lFl3WY6vkftuO0c4XNqSL3QKKvuBhF852KA8fLfhWjEjKp7hllGo7DIAtZl8r2YvLhJl9x7x-XfGfFhkGuBYXYx"
HEADERS = {
    "Authorization": "Bearer " + API_KEY,
}
wb = Workbook()
ws = wb.active
ws.title = "Germany"
row = 1
column_city = 1
column_name = 2
column_url = 3

# todo: pridat viac miest (kubo)

germany = ["Berlin", "Frankfurt", "Hamburg", "Munich", "Essen"]
restaurant_urls = [] 

for town in germany:
    offset = 0
    while offset < 1:

        parameters = {
            "offset": offset,
            "location": town,
            "limit": 3,
            "term": "Restaurants"
        }

        response = requests.get(URL, headers=HEADERS, params=parameters)
        query = response.json()["businesses"]

        for q in query:
            restaurant_urls.append(q["url"])
            ws.cell(row=row, column=column_city, value=q["location"]["city"])
            ws.cell(row=row, column=column_name, value=q["name"])
            ws.cell(row=row, column=column_url, value=q["url"])
            row += 1

        offset += 3
wb.save(filename="test-germany1.xlsx")
